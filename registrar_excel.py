
from __future__ import annotations

import json
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from flujo06 import (
    METODO_CHUNKING,
    METODO_GENERACION,
    VERSION_SALIDA,
    calcular_feature_global_desde_chunks,
    construir_entrada_06,
    normalizar_chunks_06,
    normalizar_entidades_06,
    texto_plano,
    to_json,
    valor_excel,
)

RESULTADOS_DIR = Path(__file__).resolve().parent / "resultados"
EXCEL_PATH = RESULTADOS_DIR / "registro_flujo_tesis_01_a_06.xlsx"

SHEETS = {
    "00_Indice": ["flujo", "hoja", "descripcion"],
    "01_Registro": [
        "registro_id", "fecha_hora", "diagnostico_id", "source", "sentence",
        "modelos_seleccionados", "ner_modelo", "ner_modo", "total_entidades", "total_chunks",
        "total_tokens_bio", "estado", "observacion"
    ],
    "01_BIO_Tokens": [
        "registro_id", "diagnostico_id", "token_id", "token", "bio_tag", "label_clinico", "entidad_clinica", "metodo_etiquetado"
    ],
    "02_NER_Entidades": [
        "registro_id", "diagnostico_id", "entidad_id", "texto", "label", "start_char", "end_char", "negado", "metodo_ner"
    ],
    "04_Documentos_Para_06": [
        "id", "source", "sentence", "metodo_chunking", "total_chunks", "estado_chunking",
        "entidades_ner_json", "chunks_json", "feature_engineering_global_json", "tokens_bio_json"
    ],
    "04_Chunks": [
        "registro_id", "diagnostico_id", "chunk_id", "chunk", "labels_chunk_texto", "categoria_principal",
        "start_char", "end_char", "longitud_palabras", "score_complejidad", "nivel_complejidad", "requiere_explicacion",
        "entidades_chunk_json", "feature_engineering_np"
    ],
    "06_Resultados": [
        "registro_id", "id", "source", "sentence", "tecnica_chunking", "total_chunks", "estado_chunking",
        "chunks_json_original", "entrada_llm_json_usada", "feature_engineering_global_json", "entidades_originales_ner_json",
        "resultado_llm_json", "respuesta_llm_raw", "analisis_global_json", "analisis_feature_engineering_modelo_json",
        "chunks_analizados_json", "diagnostico_simplificado", "simplificacion_chunks_usados_json", "simplificacion_chunks_descartados_json",
        "explicaciones_chunks_json", "explicacion_chunks_usados_json", "explicacion_chunks_descartados_json", "resumen_automatico",
        "resumen_chunks_usados_json", "resumen_chunks_descartados_json", "evaluacion_interna_json", "alertas_semanticas_json",
        "advertencias_clinicas_json", "llm_provider", "llm_model", "model_alias", "estado_llm", "error", "metodo_generacion",
        "version_salida", "temperature", "max_tokens", "timestamp_inicio", "timestamp_fin", "elapsed_seconds",
        "input_tokens", "output_tokens", "total_tokens", "usage_json", "raw_response_json", "resultado_modelo_json", "respuesta_modelo_raw"
    ],
    "06_Explicaciones_Chunks": [
        "registro_id", "id", "model_alias", "llm_provider", "llm_model", "chunk_id", "explicacion_simple", "terminos_clave_json"
    ],
    "06_Resumen_Modelos": [
        "registro_id", "model_alias", "llm_provider", "llm_model", "estado_llm", "cantidad"
    ],
}

INDICE_ROWS = [
    ["00", "00_Indice", "Descripción de las hojas generadas automáticamente."],
    ["01", "01_Registro", "Registro maestro por diagnóstico ingresado desde la página."],
    ["01", "01_BIO_Tokens", "Tokens y etiquetas BIO generadas para cada diagnóstico."],
    ["02", "02_NER_Entidades", "Entidades clínicas detectadas por NER."],
    ["04", "04_Documentos_Para_06", "Estructura de entrada que alimenta el flujo 06."],
    ["04", "04_Chunks", "Chunks clínicos con entidades, etiquetas y rasgos de complejidad."],
    ["06", "06_Resultados", "Salida multimodelo con simplificación, explicación, resumen y trazabilidad."],
    ["06", "06_Explicaciones_Chunks", "Explicaciones normalizadas por modelo y chunk."],
    ["06", "06_Resumen_Modelos", "Resumen incremental de modelos ejecutados por diagnóstico."],
]


def append_seguro(ws: Worksheet, values: List[Any]) -> None:
    ws.append([valor_excel(value) for value in values])


def _aplicar_estilo_hoja(ws: Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.freeze_panes = "A2"

    for col in ws.columns:
        header = str(col[0].value or "")
        letter = col[0].column_letter
        if header.endswith("json") or header in {"resultado_llm_json", "respuesta_llm_raw", "entrada_llm_json_usada", "chunks_json_original", "sentence", "diagnostico_simplificado", "resumen_automatico", "chunk", "feature_engineering_np"}:
            ws.column_dimensions[letter].width = 45
        elif header in {"registro_id", "diagnostico_id", "timestamp_inicio", "timestamp_fin"}:
            ws.column_dimensions[letter].width = 24
        else:
            ws.column_dimensions[letter].width = 18

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border


def crear_excel_si_no_existe(path: Path = EXCEL_PATH) -> Path:
    RESULTADOS_DIR.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return path
    wb = Workbook()
    ws_default = wb.active
    wb.remove(ws_default)
    for sheet_name, headers in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        if sheet_name == "00_Indice":
            for row in INDICE_ROWS:
                ws.append(row)
        _aplicar_estilo_hoja(ws)
    wb.save(path)
    return path


def _asegurar_hojas(wb) -> None:
    for sheet_name, headers in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
            if sheet_name == "00_Indice":
                for row in INDICE_ROWS:
                    ws.append(row)
            _aplicar_estilo_hoja(ws)
        else:
            ws = wb[sheet_name]
            if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
                ws.append(headers)
            _aplicar_estilo_hoja(ws)


def _parse_json_safe(value: Any, default: Any):
    if isinstance(value, (dict, list)):
        return value
    if not value:
        return default
    try:
        return json.loads(str(value))
    except Exception:
        return default


def registrar_analisis_excel(
    diagnostico_original: str,
    modelos: List[str],
    entidades: List[Dict[str, Any]],
    bio_tags: List[Dict[str, Any]],
    chunks: List[Dict[str, Any]],
    resultados_modelos: List[Dict[str, Any]],
    ner_modelo: str,
    ner_modo: str,
    diagnostico_id: Optional[str] = None,
    fuente: str = "Ingreso manual web",
    observacion: str = "",
    estado: str = "procesado",
) -> Dict[str, str]:
    path = crear_excel_si_no_existe()
    wb = load_workbook(path)
    _asegurar_hojas(wb)

    fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    registro_id = f"WEB-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
    diagnostico_id = diagnostico_id or registro_id

    entidades_06 = normalizar_entidades_06(entidades)
    chunks_06 = normalizar_chunks_06(chunks, entidades)
    feature_global = calcular_feature_global_desde_chunks(chunks_06)
    entrada_06 = construir_entrada_06(diagnostico_original, entidades, chunks)

    ws = wb["01_Registro"]
    append_seguro(ws, [
        registro_id, fecha_hora, diagnostico_id, fuente, diagnostico_original,
        ", ".join(modelos), ner_modelo, ner_modo, len(entidades_06), len(chunks_06),
        len(bio_tags), estado, observacion,
    ])

    ws_bio = wb["01_BIO_Tokens"]
    for idx, token in enumerate(bio_tags, start=1):
        bio = token.get("bio", token.get("bio_tag", "O"))
        label = bio.split("-", 1)[1] if "-" in bio else "O"
        append_seguro(ws_bio, [
            registro_id, diagnostico_id, idx, token.get("token", token.get("text", "")), bio,
            label, token.get("entidad_clinica", ""), token.get("metodo_etiquetado", ner_modo),
        ])

    ws_ner = wb["02_NER_Entidades"]
    for ent in entidades_06:
        append_seguro(ws_ner, [
            registro_id, diagnostico_id, ent.get("entidad_id"), ent.get("texto"), ent.get("label"),
            ent.get("start_char"), ent.get("end_char"), ent.get("negado"), ent.get("metodo_ner"),
        ])

    ws_doc06 = wb["04_Documentos_Para_06"]
    append_seguro(ws_doc06, [
        diagnostico_id, fuente, diagnostico_original, METODO_CHUNKING, len(chunks_06), "ok",
        to_json(entidades_06), to_json(chunks_06), to_json(feature_global), to_json(bio_tags),
    ])

    ws_chunks = wb["04_Chunks"]
    for ch in chunks_06:
        append_seguro(ws_chunks, [
            registro_id, diagnostico_id, ch.get("chunk_id"), ch.get("chunk"), ch.get("labels_chunk_texto"),
            ch.get("categoria_principal"), ch.get("start_char"), ch.get("end_char"), ch.get("longitud_palabras"),
            ch.get("score_complejidad"), ch.get("nivel_complejidad"), ch.get("requiere_explicacion"),
            to_json(ch.get("entidades_chunk_json", [])), to_json(ch.get("feature_engineering_np", {})),
        ])

    ws_resultados = wb["06_Resultados"]
    ws_expl = wb["06_Explicaciones_Chunks"]
    ws_resumen = wb["06_Resumen_Modelos"]

    for salida in resultados_modelos:
        model_alias = salida.get("model_alias", salida.get("modelo_id", ""))
        provider = salida.get("llm_provider", salida.get("proveedor", ""))
        llm_model = salida.get("llm_model", salida.get("modelo_usado", ""))
        estado_llm = salida.get("estado_llm", "ok")

        append_seguro(ws_resultados, [
            registro_id, diagnostico_id, fuente, diagnostico_original,
            METODO_CHUNKING, len(chunks_06), "ok",
            to_json(chunks_06), to_json(entrada_06), to_json(feature_global), to_json(entidades_06),
            salida.get("resultado_llm_json", salida.get("resultado_modelo_json", "")),
            salida.get("respuesta_llm_raw", salida.get("respuesta_modelo_raw", "")),
            salida.get("analisis_global_json", ""),
            salida.get("analisis_feature_engineering_modelo_json", ""),
            salida.get("chunks_analizados_json", ""),
            texto_plano(salida.get("diagnostico_simplificado", "")),
            salida.get("simplificacion_chunks_usados_json", ""),
            salida.get("simplificacion_chunks_descartados_json", ""),
            salida.get("explicaciones_chunks_json", ""),
            salida.get("explicacion_chunks_usados_json", ""),
            salida.get("explicacion_chunks_descartados_json", ""),
            texto_plano(salida.get("resumen_automatico", "")),
            salida.get("resumen_chunks_usados_json", ""),
            salida.get("resumen_chunks_descartados_json", ""),
            salida.get("evaluacion_interna_json", ""),
            salida.get("alertas_semanticas_json", ""),
            salida.get("advertencias_clinicas_json", ""),
            provider, llm_model, model_alias, estado_llm, salida.get("error", ""),
            salida.get("metodo_generacion", METODO_GENERACION),
            salida.get("version_salida", VERSION_SALIDA),
            salida.get("temperature", ""), salida.get("max_tokens", ""),
            salida.get("timestamp_inicio", ""), salida.get("timestamp_fin", ""), salida.get("elapsed_seconds", ""),
            salida.get("input_tokens", ""), salida.get("output_tokens", ""), salida.get("total_tokens", ""),
            salida.get("usage_json", ""), salida.get("raw_response_json", ""),
            salida.get("resultado_modelo_json", salida.get("resultado_llm_json", "")),
            salida.get("respuesta_modelo_raw", salida.get("respuesta_llm_raw", "")),
        ])

        explicaciones = _parse_json_safe(salida.get("explicaciones_chunks_json", "[]"), [])
        for exp in explicaciones:
            if isinstance(exp, dict):
                append_seguro(ws_expl, [
                    registro_id, diagnostico_id, model_alias, provider, llm_model, exp.get("chunk_id"),
                    exp.get("explicacion_simple", exp.get("explicacion", "")),
                    to_json(exp.get("terminos_clave", [])),
                ])

        append_seguro(ws_resumen, [
            registro_id, model_alias, provider, llm_model, estado_llm, 1,
        ])

    for sheet_name in SHEETS:
        _aplicar_estilo_hoja(wb[sheet_name])

    wb.save(path)
    return {"registro_id": registro_id, "excel_path": str(path)}
